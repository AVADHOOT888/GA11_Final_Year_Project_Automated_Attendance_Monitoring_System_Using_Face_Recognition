import os
from os import listdir
from PIL import Image as Img
from numpy import asarray
from numpy import expand_dims
from matplotlib import pyplot
import numpy as np
import pickle
import cv2
import base64
from keras_facenet import FaceNet
import openpyxl
from pytz import timezone
import pandas as pd
import datetime
from IPython.display import display, Javascript
# from google.colab.output import eval_js
from base64 import b64decode
from IPython.display import Image
# from database import database, students

MyFaceNet = FaceNet()
HaarCascade = cv2.CascadeClassifier(cv2.samples.findFile(
    cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'))
database = {}
folder='Images/'
students=[]


for filename in listdir(folder):
    path = folder + filename
    gbr1 = cv2.imread(path)
    if(gbr1 is not None):
      students.append(filename.split(".")[0])
    
    if gbr1 is None:
        print(f"Error reading image {filename}")
        continue
    
    wajah = HaarCascade.detectMultiScale(gbr1, 1.1, 4)
    if len(wajah) == 0:
        print(f"No face detected in image {filename}")
        continue
    
    x1, y1, width, height = wajah[0]
    x1, y1 = abs(x1), abs(y1)
    x2, y2 = x1 + width, y1 + height
    
    gbr = cv2.cvtColor(gbr1, cv2.COLOR_BGR2RGB)
    gbr = Img.fromarray(gbr)
    gbr_array = np.array(gbr)
    
    face = gbr_array[y1:y2, x1:x2]
    face = Img.fromarray(face)
    face = face.resize((160, 160))
    face = np.array(face)
    
    face = np.expand_dims(face, axis=0)
    signature = MyFaceNet.embeddings(face)
    
    database[os.path.splitext(filename)[0]] = signature
myfile = open("data.pkl", "wb")
pickle.dump(database, myfile)
myfile.close()

myfile = open("data.pkl", "rb")
database = pickle.load(myfile)
myfile.close()

# print(database)


import pickle
myfile = open("data.pkl", "wb")
pickle.dump(database, myfile)
myfile.close()



myfile = open("data.pkl", "rb")
database = pickle.load(myfile)
myfile.close()

# print(database)
# students=['Abhimanyu_BabarPatil', 'Avadhoot_Autade', 'Harshad_Gurav', 'Kartik_Deshmukh', 'Omkar_Waingade', 'Pratik_Rabade', 'Rahul_Pawar', 'Sanket_Patil', 'Suyog_Mokashi']
def js_to_image(js_reply):
    image_bytes = b64decode(js_reply.split(',')[1])
    jpg_as_np = np.frombuffer(image_bytes, dtype=np.uint8)
    img = cv2.imdecode(jpg_as_np, flags=1)
    return img


def generate_excel(students, present_students, filename='attendance.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.cell(row=1, column=1, value="Student")
    ws.cell(row=1, column=2, value="Present/Absent")

    row = 2

    for student in students:
        ws.cell(row=row, column=1, value=student)
        if student in present_students:
            ws.cell(row=row, column=2, value="Present")
        else:
            ws.cell(row=row, column=2, value="Absent")
        row += 1

    wb.save(filename)


# def generate_attendance(students, present_list):
#     day_of_week = datetime.datetime.today().strftime('%A')
#     # current_time = datetime.datetime.now().time()
#     current_time=datetime.datetime.now(timezone("Asia/Kolkata")).time()
#     print(current_time)

#     # Read timetable from CSV file
#     timetable = pd.read_csv('/content/gdrive/MyDrive/FinalYearProject/timetable.csv')
#     current_subject = None

#     # Determine the current subject based on the day of the week and time
#     schedule = timetable.loc[timetable['Day'] == day_of_week]
#     print(schedule)
#     for i, row in schedule.iterrows():
#         start_time = datetime.datetime.strptime(row['Timing'].split(' - ')[0], '%I:%M %p').time()
#         end_time = datetime.datetime.strptime(row['Timing'].split(' - ')[1], '%I:%M %p').time()
#         if start_time <= current_time <= end_time:
#             current_subject = row['Subject']
#             break

#     if current_subject is None:
#         print('No class is currently in session.')
#         return

#     # Generate filename based on current date and subject
#     date_today = datetime.date.today().strftime('%d-%m-%Y')
#     filename = f'attendance_{current_subject}_{date_today}.xlsx'

#     attendance = pd.DataFrame(columns=['Student Name', 'Present', 'Absent', 'Timing'])

#     for student in students:
#         if student == 'unknown':
#             continue
#         else:
#           present = 1 if student in present_list else 0
#           absent = 1 - present
#         row = pd.DataFrame([[student, present, absent, current_time]], columns=['Student Name', 'Present', 'Absent', 'Timing'])
#         attendance = attendance.append(row, ignore_index=True)

#     attendance.to_excel(filename, index=False)
#     print(f"Attendance sheet generated: {filename}")
#     print("Students: ",students)
#     print("present_list",present_list)

# # generate_attendance(students, present_list)

def findFaces(data, threshold=1):
    gbr1 = js_to_image(data)
    gbr = cv2.cvtColor(gbr1, cv2.COLOR_BGR2RGB)
    gbr = Img.fromarray(gbr)                  # Convert from OpenCV to PIL
    gbr_array = asarray(gbr)

    wajah = HaarCascade.detectMultiScale(gbr1, 1.1, 4)
    present_Students = []
    accuracies = []
    for (x1, y1, w, h) in wajah:
        x1, y1 = abs(x1), abs(y1)
        x2, y2 = x1 + w, y1 + h

        face = gbr_array[y1:y2, x1:x2]

        face = Img.fromarray(face)
        face = face.resize((160, 160))
        face = asarray(face)

        face = expand_dims(face, axis=0)
        signature = MyFaceNet.embeddings(face)

        min_dist = 100
        identity = 'unknown'
        for key, value in database.items():
            dist = np.linalg.norm(value - signature)
            if dist < min_dist:
                min_dist = dist
                identity = key
            if min_dist > threshold:
                identity = 'unknown'
        # calculate the accuracy of the recognized face
        accuracy = (threshold - min_dist) / threshold
        accuracies.append(accuracy)

        present_Students.append(identity)

        cv2.putText(gbr1, identity, (x1, y1), cv2.FONT_HERSHEY_SIMPLEX,
                    0.5, (255, 255, 0), 1, cv2.LINE_AA)
        cv2.rectangle(gbr1, (x1, y1), (x2, y2), (0, 255, 0), 2)

    filename = 'photo.jpg'
    cv2.imwrite(filename, gbr1)
    print(present_Students)
    generate_excel(students, present_Students)  # code without schedule
    # generate_attendance(students, present_Students)  Main code with schedule

    return filename


def findFacesFromLocalFile(file_path):
    with open(file_path, "rb") as image_file:
        encoded_string = base64.b64encode(image_file.read()).decode("utf-8")
    data = "data:image/jpeg;base64," + encoded_string
    return findFaces(data)


####Main Code######
def take_photo():
    filename = findFacesFromLocalFile("Test/Birthday.jpeg")
    return filename
##################

# def take_photo(img):
#   filename = findFacesFromLocalFile(img)
#   return filename


try:
    filename = take_photo()
    print('Saved to {}'.format(filename))

    display(Image(filename))
except Exception as err:
    print(str(err))
